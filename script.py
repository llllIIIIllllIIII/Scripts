from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import os
import time
# from selenium.webdriver.common.action_chains import ActionChains
from openpyxl import load_workbook
import keyboard


url = 'https://proteacher.moe.edu.tw/#login'
year_list = list(range(95, 111))

ACCOUNT = 'utcollege'
PASSWORD = 'i-love-tim99'

Wait_time = 5


try:
    driver = webdriver.Chrome()
except Exception as e:
    print(e)
    os.system('pause')

driver.get(url)

acc = (By.XPATH,('//*[@id="root"]/div/div[3]/div/div/div/div/div/div[2]/div/div/div[2]/div[2]/input'))
acc_input = WebDriverWait(driver,Wait_time).until(EC.presence_of_element_located(acc),"找不到指定的元素")
acc_input.send_keys(ACCOUNT)

pwd = (By.XPATH,'//*[@id="root"]/div/div[3]/div/div/div/div/div/div[2]/div/div/div[2]/div[3]/input')
pwd_input = WebDriverWait(driver,Wait_time).until(EC.presence_of_element_located(pwd),"找不到指定的元素")
pwd_input.send_keys(PASSWORD)

submit = (By.XPATH,'//*[@id="root"]/div/div[3]/div/div/div/div/div/div[2]/div/div/div[2]/div[4]/button')
sub_input = WebDriverWait(driver,Wait_time).until(EC.presence_of_element_located(submit),"找不到指定的元素")
sub_input.click()
time.sleep(3)
driver.refresh()
while True:
    try:
        file_name = str(input("輸入檔案名稱(確認為副檔名為.xlsx之excel檔案 不須輸入副檔名):"))
        file_name = file_name + '.xlsx'
        wb = load_workbook(file_name)
        break
    except Exception as e:
        print(e)
        print("請確認檔案名稱輸入正確或檢查此檔案是否存在")
    
ws = wb['工作表1']

while True:
    try:
        col_name = str(input("輸入需要之欄位(ex:A,B,C):"))
        break
    except Exception as e:
        print(e)
        print("請確認檔案名稱輸入正確或檢查此檔案是否存在")

ID_list = []
for tid in ws[col_name]:
    ID_list.append(tid.value)
print (ID_list)

for i in ID_list:
    d_url = 'https://proteacher.moe.edu.tw/#function/social/schoolform'
    driver.get(d_url)
    name = (By.XPATH,('//*[@id="root"]/div/div[2]/div/div/div/div/div/span/span/input'))
    name_input = WebDriverWait(driver,Wait_time).until(EC.presence_of_element_located(name),"找不到輸入的框框")
    name_input.send_keys(i)
    searchBtn = (By.XPATH,'//*[@id="root"]/div/div[2]/div/div/div/div/div/span/span/span/button')
    searchBtn_A = WebDriverWait(driver,Wait_time).until(EC.presence_of_element_located(searchBtn),"找不到指定的元素")
    searchBtn_A.click()
    keyboard.wait("Scroll Lock")


        


os.system('pause')
# driver.quit()