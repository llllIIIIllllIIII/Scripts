import glob
from imaplib import Int2AP
import PyPDF2
from openpyxl import load_workbook

y01 = glob.glob('Y01*pdf')
y02 = glob.glob('Y02*pdf')
fe = '.pdf'
savepath = 'D:\\Pythontest\\較專\\OneDrive_1_2022-3-25\\0325\\'
filepath = 'D:\\Pythontest\\較專\\OneDrive_1_2022-3-25\\S3排序_重複標記.xlsx'

wb = load_workbook(filepath)

ws = wb['1110325台中提供名單']
Filename =[i.value for i in (list(ws['c']))] 
# print(Filename)

for i in range(0,len(y01)):
    output = ""
    mergeFile = PyPDF2.PdfFileMerger()
    id = y01[i][4:9]
    print(y02[i])
    mergeFile.append(PyPDF2.PdfFileReader(y01[i], 'rb'))
    mergeFile.append(PyPDF2.PdfFileReader(y02[i], 'rb'))
    for name in Filename:
        if id in name:
            output = name
            Filename.remove(name)
            break
    output = savepath + output + fe
    print(output+"\n"+id)
    # input()
    mergeFile.write(output)
    mergeFile.close()
print(Filename)
    
# print('abc')