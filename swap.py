import tkinter
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from tkinter import *
from tkinter import  filedialog,messagebox

order = ('案件編號', '表單群稱', '表單名稱', '表單資料', '年度', '縣市', '學校名稱', '老師姓名',
         '流程名稱', '表單創建時間', '表單最後更新時間', '社群類型-辦理-是否續辦', '召集人教授科目', '社群類型-學校', '社群類型-是否申請經費', '社群類型-是否邀請專家諮詢',
         '召集人是否曾參與召集人培訓', '社群成員-社群人數(含召集人）', '社群成員-欄數', '進度規劃（視各縣市政府規定進行規劃）-場次', '社群級別(縣市設定之級別)*', '社群類型-學年 學科/領域',
         '專業學習社群內容 (可複選)', '召集人姓名', '召集人職稱', '社群欲培養學生之核心素養(可複選)-B.溝通互動', '社群欲培養學生之核心素養(可複選)-C.社會參與',
         '社群類型-任教階段', '召集人聯絡電話', '社群欲培養學生之核心素養(可複選)-A.自主行動', '社群運作目標(可複選)', '社群進行方式(可複選)',
         '社群簡介(100~200字以內)*', '社群名稱(至多20字)*', '學校名稱*', '召集人e-mail', 'upload')
order2 = ('案件編號', '表單群稱', '表單名稱', '表單資料', '年度', '縣市', '學校名稱', '老師姓名',

         '流程名稱', '表單創建時間', '表單最後更新時間','1.外聘講座鐘點費-總價', '1.外聘講座鐘點費-單價', '1.外聘講座鐘點費-數量', '1.外聘講座鐘點費-總價', '1.外聘講座鐘點費-單價', '1.外聘講座鐘點費-數量',
          '2.內聘講座鐘點費-總價',
          '2.內聘講座鐘點費-單價',
          '2.內聘講座鐘點費-數量',
          '3.外聘諮詢費-總價',
          '3.外聘諮詢費-單價',
          '3.外聘諮詢費-數量',
          '4.二代健保補充保費-總價',
          '4.二代健保補充保費-單價',
          '4.二代健保補充保費-數量',
          '5.講座交通費-總價',
          '5.講座交通費-單價',
          '5.講座交通費-數量',
          '6.印刷費-總價',
          '6.印刷費-單價',
          '6.印刷費-數量',
          '7.教材教具費-總價',
          '7.教材教具費-單價',
          '7.教材教具費-數量',
          '8.膳費-總價',
          '8.膳費-單價',
          '8.膳費-數量',
          '9.雜支-總價',
          '9.雜支-單價',
          '9.雜支-數量',
          '總計(新台幣元整)',
          '隸屬縣市行政區',
          '學校名稱',
          '老師姓名')

filepath = ""

def main(filepath):
    if not filepath:
        return
    while True:
        try:
            wb = load_workbook(filepath)
            break
        except Exception as e:
            print(e)
    ws = wb.worksheets[0]


    for index, name in enumerate(ws['1']):
        if name.value == order[index]:
            continue
        for i in range(index,len(ws['1'])):
            if  ws[get_column_letter(i+1)][0].value == order[index]:
                for k in range(len(ws[get_column_letter(i+1)])):
                    ws[get_column_letter(index+1)][k].value, ws[get_column_letter(
                        i+1)][k].value = ws[get_column_letter(i+1)][k].value, ws[get_column_letter(index+1)][k].value
                continue        

    for index, name in enumerate(ws['1']):
        if name.value == order[index]:
            print('True')
        else:
            print(order[index] + "False")
    wb.save(filepath)
    messagebox.showinfo("提示", "排序完成")


            




def getFilePath():
    global filepath
    filepath = filedialog.askopenfilename(title ="開啟檔案",filetype=[("Excel","*.xlsx")])
    main(filepath)
    


if __name__ == '__main__':
    window = Tk()
    window.title("自動排序")
    window.geometry("500x500")
    window.resizable(width=0, height=0)
    pixelVirtual = tkinter.PhotoImage(width=1, height=1)
    button = Button(text="開啟檔案",command=getFilePath,image=pixelVirtual,width=150,height=50,compound="c")
    button.pack()
    window.mainloop()