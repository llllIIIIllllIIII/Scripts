from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import os
import time
import json
import urllib.parse

durl = urllib.parse.unquote(
    '/api/v1/result/type/social/Y04/file?year=111\u0026schoolId=')
durl = 'https://proteacher.moe.edu.tw/'+durl


searchList = []
filenameList = []
finishList = []
ACCOUNT = 'taichungcityCoach'
PASSWORD = 'Taichung407@'
filepath = r"C:\Users\henry\Downloads\校名.xlsx"

try:
    wb = load_workbook(filepath)
except Exception as e:
    print(e)

ws = wb.worksheets[1]

filenameList = [i.value for i in ws['a'][1:]]
searchList = [i.value for i in ws['o'][1:]]


Wait_time = 5
anchor = None
data = []
try:
    options = webdriver.ChromeOptions()
    # options.add_experimental_option('prefs', {
    #     # Change default directory for downloads
    #     "download.default_directory": "X:\python\教專\學校檔案",
    #     "download.prompt_for_download": False,  # To auto download the file
    #     "download.directory_upgrade": True,
    #     # It will not show PDF directly in chrome
    #     "plugins.always_open_pdf_externally": True
    # })
    driver = webdriver.Chrome()
except Exception as e:
    print(e)
    os.system('pause')

driver.get('https://proteacher.moe.edu.tw/#login')

acc = (By.XPATH,
       ('//*[@id="root"]/div/div[3]/div/div/div/div/div/div[2]/div/div/div[2]/div[2]/input'))
acc_input = WebDriverWait(driver, Wait_time).until(
    EC.presence_of_element_located(acc), "找不到指定的元素")
acc_input.send_keys(ACCOUNT)

pwd = (By.XPATH,
       '//*[@id="root"]/div/div[3]/div/div/div/div/div/div[2]/div/div/div[2]/div[3]/input')
pwd_input = WebDriverWait(driver, Wait_time).until(
    EC.presence_of_element_located(pwd), "找不到指定的元素")
pwd_input.send_keys(PASSWORD)

submit = (
    By.XPATH, '//*[@id="root"]/div/div[3]/div/div/div/div/div/div[2]/div/div/div[2]/div[4]/button')
sub_input = WebDriverWait(driver, Wait_time).until(
    EC.presence_of_element_located(submit), "找不到指定的元素")
sub_input.click()
time.sleep(3)
driver.refresh()

for id, search in enumerate(searchList):
    if search in finishList:
        continue
    finishList.append(search)
    driver.get(
        'https://proteacher.moe.edu.tw/api/v1/flowResults/public/social?keyword='+search+'&year=111')
    driver.implicitly_wait(1)
    pre = driver.find_element_by_tag_name("pre").text

    data = json.loads(pre)
    if data['data'] == None:
        print(data)
        input()
        continue
    schoolId = data['data'][0]['schoolId']
    

    driver.get(durl+str(schoolId))
